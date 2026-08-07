"""Uji pembuatan dan pembacaan template universal."""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from inaproc_autoinput import references as ref
from inaproc_autoinput import workbook
from inaproc_autoinput.schema import all_fields, attribute_pairs
from inaproc_autoinput.workbook import FIRST_DATA_ROW, ROW_LABEL, SHEET_PRODUK


def test_template_punya_semua_kolom(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    ws = load_workbook(path)[SHEET_PRODUK]

    labels = [
        str(ws.cell(ROW_LABEL, i).value or "").strip()
        for i in range(1, ws.max_column + 1)
    ]
    assert labels[:2] == ["Kategori", "Tipe Produk"]
    assert len(labels) == len(all_fields())
    assert "Nama Produk" in labels
    assert "Atribut 1" in labels and "Nilai 1" in labels
    # Berkas dipilih di aplikasi, bukan diketik di Excel.
    assert not [l for l in labels if l.startswith(("Foto", "Dokumen", "Berkas"))]
    assert "Video Produk" not in labels
    assert "Berat Produk (gram)" in labels  # khusus kategori Barang


def test_template_menandai_kolom_wajib(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    ws = load_workbook(path)[SHEET_PRODUK]
    keterangan = {
        str(ws.cell(ROW_LABEL, i).value): str(ws.cell(3, i).value)
        for i in range(1, ws.max_column + 1)
    }
    assert keterangan["Nama Produk"] == "(Wajib)"
    assert keterangan["Tipe Produk"] == "(Opsional)"  # portal menentukannya sendiri


def test_kolom_berpilihan_punya_dropdown(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    ws = load_workbook(path)[SHEET_PRODUK]

    rumus = {}
    for dv in ws.data_validations.dataValidation:
        for rng in str(dv.sqref).split():
            rumus[rng.split(":")[0].rstrip("0123456789")] = dv.formula1

    kolom = {
        str(ws.cell(ROW_LABEL, i).value or "").strip(): get_column_letter(i)
        for i in range(1, ws.max_column + 1)
    }
    # Daftar yang muat ditulis langsung di sel.
    assert rumus[kolom["PPN"]] == '"0%,12%"'
    assert rumus[kolom["Klasifikasi Produk"]] == '"Lokal,Import"'
    assert rumus[kolom["Lokasi Produksi"]].startswith('"Diproduksi di seluruh')
    # Yang melewati batas 255 karakter menunjuk sheet rujukan.
    assert workbook.SHEET_PILIHAN in rumus[kolom["Satuan Produk"]]
    assert len('"{}"'.format(",".join(ref.SATUAN_PRODUK))) > 255


def test_sheet_pilihan_berisi_daftar_panjang(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    ws = load_workbook(path)[workbook.SHEET_PILIHAN]
    judul = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    assert "Satuan Produk" in judul

    kolom = judul.index("Satuan Produk") + 1
    nilai = {ws.cell(r, kolom).value for r in range(2, ws.max_row + 1)}
    assert "Meter" in nilai and "M³" in nilai
    assert len(nilai - {None}) == 82


def test_kategori_dropdown_menunjuk_daftar_kategori(tmp_path, monkeypatch):
    from inaproc_autoinput.categories import Catalog, Category

    katalog = Catalog([
        Category(id="a", level=1, name="Bidang Bina Marga", product_type="SERVICE",
                 children=(Category(id="b", level=2, name="Divisi 3",
                                    product_type="SERVICE",
                                    children=(Category(id="c", level=3,
                                                       name="3.1 Galian",
                                                       product_type="SERVICE"),)),)),
    ])
    path = workbook.create_template(tmp_path / "template.xlsx", catalog=katalog)
    wb = load_workbook(path)
    ws = wb[SHEET_PRODUK]

    kolom = next(i for i in range(1, ws.max_column + 1)
                 if str(ws.cell(ROW_LABEL, i).value).strip() == "Kategori")
    huruf = get_column_letter(kolom)
    rumus = next(dv.formula1 for dv in ws.data_validations.dataValidation
                 if str(dv.sqref).startswith(huruf))
    assert workbook.SHEET_KATEGORI in rumus

    # Kolom E berisi jalur lengkap yang jadi sumber dropdown.
    kategori = wb[workbook.SHEET_KATEGORI]
    assert kategori.cell(1, 5).value == "Jalur Kategori"
    assert kategori.cell(2, 5).value == "Bidang Bina Marga > Divisi 3 > 3.1 Galian"


def test_template_punya_sheet_petunjuk(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    wb = load_workbook(path)
    assert workbook.SHEET_PETUNJUK in wb.sheetnames
    teks = "\n".join(
        str(c.value) for row in wb[workbook.SHEET_PETUNJUK].iter_rows() for c in row
    )
    assert "tab 'Berkas' di aplikasi" in teks  # berkas tidak lagi di Excel


def _isi(path, baris: list[dict]):
    """Tulis beberapa baris ke template, dicocokkan lewat judul kolomnya."""
    wb = load_workbook(path)
    ws = wb[SHEET_PRODUK]
    kolom = {
        str(ws.cell(ROW_LABEL, i).value or "").strip(): i
        for i in range(1, ws.max_column + 1)
    }
    for offset, data in enumerate(baris):
        for label, value in data.items():
            ws.cell(FIRST_DATA_ROW + offset, kolom[label], value)
    wb.save(path)


def test_baca_kembali_yang_sudah_diisi(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    _isi(path, [
        {
            "Kategori": "Bidang Bina Marga > Divisi 3 > 3.1 Galian",
            "Nama Produk": "Galian Biasa untuk Badan Jalan",
            "Harga Produk": 85000,
            "Atribut 1": "Satuan Pengukuran",
            "Nilai 1": "M3",
        },
        {"Nama Produk": "Timbunan Pilihan dari Sumber Galian"},
    ])

    rows = workbook.read_workbook(path)
    assert len(rows) == 2
    assert rows[0]["_row"] == FIRST_DATA_ROW
    assert rows[0]["nama_produk"] == "Galian Biasa untuk Badan Jalan"
    assert rows[0]["harga_produk"] == "85000"
    assert attribute_pairs(rows[0]) == {"Satuan Pengukuran": "M3"}
    assert rows[1]["_row"] == FIRST_DATA_ROW + 1


def test_baris_kosong_dilewati(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    _isi(path, [{"Nama Produk": "Produk Pertama Sekali"}])
    wb = load_workbook(path)
    wb[SHEET_PRODUK].cell(FIRST_DATA_ROW + 5, 5, None)  # sentuh sel jauh di bawah
    wb.save(path)

    assert len(workbook.read_workbook(path)) == 1


def test_kolom_boleh_digeser(tmp_path):
    """Pembacaan mengikuti judul kolom, bukan posisinya."""
    path = workbook.create_template(tmp_path / "template.xlsx")
    wb = load_workbook(path)
    ws = wb[SHEET_PRODUK]
    ws.insert_cols(1)
    ws.cell(ROW_LABEL, 1, "Catatan Internal")
    wb.save(path)

    _isi(path, [{"Nama Produk": "Produk Setelah Kolom Digeser"}])
    rows = workbook.read_workbook(path)
    assert rows[0]["nama_produk"] == "Produk Setelah Kolom Digeser"


def test_missing_columns_melaporkan_yang_hilang(tmp_path):
    path = workbook.create_template(tmp_path / "template.xlsx")
    assert workbook.missing_columns(path) == []

    wb = load_workbook(path)
    ws = wb[SHEET_PRODUK]
    kolom = next(
        i for i in range(1, ws.max_column + 1)
        if str(ws.cell(ROW_LABEL, i).value).strip() == "Harga Produk"
    )
    ws.delete_cols(kolom)
    wb.save(path)

    assert "Harga Produk" in workbook.missing_columns(path)
