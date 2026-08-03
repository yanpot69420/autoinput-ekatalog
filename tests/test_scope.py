"""Uji penyaringan bidang: hanya konstruksi dan pertanian yang masuk template."""

from __future__ import annotations

import json

import pytest

from inaproc_autoinput import workbook
from inaproc_autoinput.categories import (
    BIDANG_DEFAULT,
    Catalog,
    Category,
    load_scope,
    save_scope,
)
from inaproc_autoinput.validation import validate


def _l1(nama: str, tipe: str = "SERVICE") -> Category:
    return Category(
        id=f"id-{nama}", level=1, name=nama, product_type=tipe,
        children=(Category(
            id=f"l2-{nama}", level=2, name="Divisi", product_type=tipe,
            children=(Category(id=f"l3-{nama}", level=3, name="Pekerjaan",
                               product_type=tipe),),
        ),),
    )


@pytest.fixture
def catalog():
    return Catalog([
        _l1("Bidang Bina Marga"),
        _l1("Tanaman dan Sarana Pendukung", "PHYSICAL"),
        _l1("Perikanan", "PHYSICAL"),        # sengaja dikeluarkan
        _l1("Makanan dan Minuman", "PHYSICAL"),
    ])


def test_bidang_default_berisi_konstruksi_dan_pertanian():
    assert len(BIDANG_DEFAULT) == 21
    assert "Bidang Bina Marga" in BIDANG_DEFAULT
    assert "Bidang Sumber Daya Air" in BIDANG_DEFAULT
    assert "Pekerjaan Cetak Sawah" in BIDANG_DEFAULT
    assert "Tanaman dan Sarana Pendukung" in BIDANG_DEFAULT


def test_perikanan_tidak_termasuk():
    assert "Perikanan" not in BIDANG_DEFAULT


def test_restrict_menyaring_bidang(catalog):
    terbatas = catalog.restrict()
    nama = [c.name for c in terbatas.roots]
    assert nama == ["Bidang Bina Marga", "Tanaman dan Sarana Pendukung"]
    assert terbatas.tally()["level3"] == 2


def test_restrict_tidak_mengubah_katalog_asli(catalog):
    catalog.restrict()
    assert catalog.tally()["level1"] == 4


def test_in_scope(catalog):
    assert catalog.in_scope("Bidang Bina Marga")
    assert not catalog.in_scope("Perikanan")
    assert catalog.in_scope("bidang bina marga")  # tidak peduli huruf besar/kecil


def test_kategori_di_luar_bidang_hanya_peringatan(catalog):
    """Kategorinya sah di portal, jadi tidak boleh diblokir -- cukup diingatkan."""
    row = {"kategori": "Perikanan > Divisi > Pekerjaan"}
    kategori = [i for i in validate(row, catalog=catalog) if i.key == "kategori"]
    assert kategori and not kategori[0].blocking
    assert "di luar bidang yang dilayani" in kategori[0].message


def test_kategori_dalam_bidang_tanpa_peringatan(catalog):
    row = {"kategori": "Bidang Bina Marga > Divisi > Pekerjaan"}
    assert not [i for i in validate(row, catalog=catalog) if i.key == "kategori"]


def test_kategori_asing_tetap_error(catalog):
    row = {"kategori": "Bidang Karangan > Divisi > Pekerjaan"}
    kategori = [i for i in validate(row, catalog=catalog) if i.key == "kategori"]
    assert kategori and kategori[0].blocking


# --- berkas pengaturan bidang ----------------------------------------------


def test_scope_bawaan_dipakai_bila_berkas_tidak_ada(tmp_path):
    assert load_scope(tmp_path / "tidak-ada.json") == BIDANG_DEFAULT


def test_scope_bisa_ditimpa_lewat_berkas(tmp_path):
    path = tmp_path / "bidang.json"
    save_scope(("Bidang Cipta Karya",), path)
    assert load_scope(path) == ("Bidang Cipta Karya",)


def test_scope_rusak_kembali_ke_bawaan(tmp_path):
    path = tmp_path / "bidang.json"
    path.write_text("{bukan json", encoding="utf-8")
    assert load_scope(path) == BIDANG_DEFAULT


def test_scope_kosong_kembali_ke_bawaan(tmp_path):
    path = tmp_path / "bidang.json"
    path.write_text(json.dumps({"bidang": []}), encoding="utf-8")
    assert load_scope(path) == BIDANG_DEFAULT


# --- template ---------------------------------------------------------------


def test_template_hanya_memuat_bidang_terpilih(tmp_path, catalog):
    from openpyxl import load_workbook

    path = workbook.create_template(
        tmp_path / "t.xlsx", catalog=catalog.restrict()
    )
    ws = load_workbook(path)[workbook.SHEET_KATEGORI]
    bidang = {ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
    assert bidang == {"Bidang Bina Marga", "Tanaman dan Sarana Pendukung"}
    assert "Perikanan" not in bidang
