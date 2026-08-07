"""Uji pohon kategori: penelusuran, pencocokan nama, dan validasi baris."""

from __future__ import annotations

import pytest

from inaproc_autoinput.categories import Catalog, Category, load_cache, save_cache
from inaproc_autoinput.validation import validate

POHON = [
    Category(
        id="l1-jasa", level=1, name="Bidang Bina Marga", product_type="SERVICE",
        children=(
            Category(
                id="l2-div3", level=2, name="Divisi 3 Pekerjaan Tanah dan Geosintetik",
                product_type="SERVICE",
                children=(
                    Category(id="l3-galian", level=3, name="3.1 Galian",
                             product_type="SERVICE"),
                    Category(id="l3-timbunan", level=3, name="3.2 Timbunan",
                             product_type="SERVICE"),
                ),
            ),
        ),
    ),
    Category(
        id="l1-barang", level=1, name="Peralatan Kantor", product_type="PHYSICAL",
        children=(
            Category(
                id="l2-meja", level=2, name="Meja", product_type="PHYSICAL",
                children=(Category(id="l3-meja", level=3, name="Meja Kerja",
                                   product_type="PHYSICAL"),),
            ),
        ),
    ),
]


@pytest.fixture
def catalog():
    return Catalog(POHON, diunduh="2026-08-03T20:00:00")


def test_tally(catalog):
    assert catalog.tally() == {"level1": 2, "level2": 2, "level3": 3}


def test_tipe_produk_diterjemahkan(catalog):
    assert catalog.find_level1("Bidang Bina Marga").tipe_produk == "Jasa"
    assert catalog.find_level1("Peralatan Kantor").tipe_produk == "Barang"


def test_level1_bisa_disaring_per_tipe(catalog):
    assert catalog.level1_names("Jasa") == ["Bidang Bina Marga"]
    assert catalog.level1_names("Barang") == ["Peralatan Kantor"]


def test_resolve_berhasil(catalog):
    node, pesan = catalog.resolve(
        "Bidang Bina Marga", "Divisi 3 Pekerjaan Tanah dan Geosintetik", "3.1 Galian"
    )
    assert node is not None and node.id == "l3-galian" and pesan == ""


def test_resolve_tidak_peduli_besar_kecil_huruf(catalog):
    node, _ = catalog.resolve(
        "bidang bina marga", "DIVISI 3 PEKERJAAN TANAH DAN GEOSINTETIK", " 3.1 galian "
    )
    assert node is not None


def test_resolve_menyarankan_ejaan_terdekat(catalog):
    _, pesan = catalog.resolve("Bidang Bina Margaa", "x", "y")
    assert "Maksudnya 'Bidang Bina Marga'?" in pesan


def test_resolve_menyebut_level_yang_salah(catalog):
    _, pesan = catalog.resolve("Bidang Bina Marga", "Divisi Karangan", "3.1 Galian")
    assert pesan.startswith("Kategori Level 2")


def test_resolve_kategori_kosong(catalog):
    _, pesan = catalog.resolve("", "", "")
    assert pesan == "Kategori Level 1 belum diisi"


def test_flatten(catalog):
    rows = catalog.flatten()
    assert len(rows) == 3
    assert rows[0] == {
        "tipe_produk": "Jasa",
        "kategori_1": "Bidang Bina Marga",
        "kategori_2": "Divisi 3 Pekerjaan Tanah dan Geosintetik",
        "kategori_3": "3.1 Galian",
        "id_kategori_3": "l3-galian",
    }


def test_cache_bolak_balik(tmp_path):
    path = tmp_path / "kategori.json"
    save_cache(POHON, path)
    roots, diunduh = load_cache(path)
    assert diunduh
    assert Catalog(roots).tally() == {"level1": 2, "level2": 2, "level3": 3}
    assert roots[0].children[0].children[0].name == "3.1 Galian"


def test_cache_rusak_tidak_bikin_error(tmp_path):
    path = tmp_path / "rusak.json"
    path.write_text("{bukan json", encoding="utf-8")
    assert load_cache(path) == ([], "")


def test_cache_hilang_tidak_bikin_error(tmp_path):
    assert load_cache(tmp_path / "tidak-ada.json") == ([], "")


# --- validasi baris ---------------------------------------------------------

BARIS = {
    "kategori": "Bidang Bina Marga > Divisi 3 Pekerjaan Tanah dan Geosintetik > 3.1 Galian",
}


def _kategori_issues(row, catalog):
    return [
        i for i in validate(row, catalog=catalog)
        if i.key in ("kategori", "tipe_produk") and i.blocking
    ]


def test_kategori_benar_lolos(catalog):
    assert _kategori_issues(BARIS, catalog) == []


def test_kategori_salah_ketik_ditolak(catalog):
    issues = _kategori_issues(dict(BARIS, kategori="Bidang Bina Marga > Divisi 3 Pekerjaan Tanah dan Geosintetik > 3.1 Galiann"), catalog)
    assert issues and "Kategori Level 3" in issues[0].message


def test_tipe_produk_diambil_dari_katalog(catalog):
    """Tidak ada kolom Tipe Produk; portal yang menentukannya dari kategori."""
    from inaproc_autoinput.validation import _check_category

    _, tipe = _check_category(BARIS, catalog)
    assert tipe == "Jasa"
    _, tipe = _check_category({"kategori": BARANG}, catalog)
    assert tipe == "Barang"


def test_resolve_path_dari_satu_kolom(catalog):
    node, pesan = catalog.resolve_path(
        "Bidang Bina Marga > Divisi 3 Pekerjaan Tanah dan Geosintetik > 3.1 Galian"
    )
    assert node is not None and node.id == "l3-galian" and pesan == ""


def test_resolve_path_kurang_lengkap(catalog):
    _, pesan = catalog.resolve_path("Bidang Bina Marga > Divisi 3")
    assert "sampai Level 3" in pesan


def test_resolve_path_kosong(catalog):
    _, pesan = catalog.resolve_path("   ")
    assert pesan == "Kategori belum diisi"


# --- kolom khusus Barang ----------------------------------------------------

BARANG = "Peralatan Kantor > Meja > Meja Kerja"


def test_berat_wajib_untuk_kategori_barang(catalog):
    """Tipe diambil dari kategori di portal, bukan dari kolom Tipe Produk."""
    issues = validate({"kategori": BARANG}, catalog=catalog)
    assert any(i.key == "berat_gram" and i.blocking for i in issues)


def test_berat_tidak_diminta_untuk_kategori_jasa(catalog):
    issues = validate(BARIS, catalog=catalog)
    assert not [i for i in issues if i.key == "berat_gram"]


def test_kolom_barang_diisi_pada_kategori_jasa_jadi_peringatan(catalog):
    issues = validate(dict(BARIS, berat_gram="1000"), catalog=catalog)
    berat = [i for i in issues if i.key == "berat_gram"]
    assert berat and not berat[0].blocking
    assert "tipe Jasa" in berat[0].message


def test_tanpa_katalog_kategori_tidak_diperiksa():
    """Aplikasi tetap jalan bila daftar kategori belum diunduh."""
    issues = validate(dict(BARIS, kategori="Karangan Bebas"), catalog=Catalog([]))
    assert not [i for i in issues if i.key == "kategori"]


def test_template_menyertakan_sheet_kategori(tmp_path, catalog):
    from openpyxl import load_workbook

    from inaproc_autoinput import workbook

    path = workbook.create_template(tmp_path / "t.xlsx", catalog=catalog)
    wb = load_workbook(path)
    assert workbook.SHEET_KATEGORI in wb.sheetnames
    ws = wb[workbook.SHEET_KATEGORI]
    assert ws.cell(1, 2).value == "Kategori Level 1"
    assert ws.max_row == 4  # 1 judul + 3 kategori level 3
