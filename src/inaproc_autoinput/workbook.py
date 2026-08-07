"""Membuat template Excel kosong dan membaca template yang sudah diisi penyedia."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from . import references as ref
from .schema import Field, all_fields, join_category

SHEET_PRODUK = "Produk"
SHEET_KATEGORI = "Daftar Kategori"
SHEET_PILIHAN = "Daftar Pilihan"
SHEET_PETUNJUK = "Petunjuk"

ROW_GROUP = 1
ROW_LABEL = 2
ROW_REQUIREMENT = 3
ROW_HINT = 4
FIRST_DATA_ROW = 5

_MAX_VALIDATED_ROW = 2000
# Batas Excel untuk daftar pilihan yang ditulis langsung di sel.
_BATAS_DAFTAR_INLINE = 255

_FILL_GROUP = PatternFill("solid", fgColor="1F3864")
_FILL_LABEL = PatternFill("solid", fgColor="2E5FA3")
_FILL_WAJIB = PatternFill("solid", fgColor="FCE4D6")
_FILL_OPSIONAL = PatternFill("solid", fgColor="F2F2F2")

PETUNJUK = [
    "CARA MENGISI TEMPLATE INI",
    "",
    "1. Isi mulai baris 5. Satu baris = satu produk.",
    "2. Baris 1 sampai 4 adalah judul dan petunjuk. Jangan dihapus atau digeser.",
    "3. Kolom bertanda (Wajib) harus diisi. Baris yang belum lengkap ditandai",
    "   merah di aplikasi dan tidak akan dijalankan.",
    "4. Kolom yang punya tanda panah: klik selnya, lalu pilih dari daftar.",
    "",
    "KATEGORI",
    "Satu kolom saja, berisi jalur lengkap dengan pemisah ' > ':",
    "  Bidang Bina Marga > Divisi 3 Pekerjaan Tanah dan Geosintetik > 3.1 Galian",
    "Bentuk ini mengikuti pemilih kategori di portal, yang juga satu kotak",
    "pencarian bertingkat, bukan tiga dropdown terpisah.",
    "",
    "Cara termudah: buka sheet 'Daftar Kategori', pakai filter untuk mencari,",
    "lalu salin isi kolom 'Jalur Kategori'. Kolom Kategori juga punya dropdown,",
    "tapi isinya ribuan baris — menyalin biasanya lebih cepat.",
    "",
    "Tipe produk tidak ditanyakan di sini — portal menentukannya sendiri dari",
    "kategori, dan aplikasi membacanya dari sana.",
    "",
    "FOTO, VIDEO, DAN DOKUMEN — TIDAK ADA DI SINI",
    "Semuanya dipilih lewat tab 'Berkas' di aplikasi, bukan diketik di Excel.",
    "Dokumen SBU dan sertifikat standar biasanya sama untuk semua produk,",
    "jadi cukup dipilih sekali dan berlaku untuk seluruh baris.",
    "Foto juga bisa dipasang sekali untuk semua baris, atau khusus per baris",
    "bila ada produk yang punya fotonya sendiri.",
    "",
    "HARGA",
    "Harga nett: belum termasuk ongkir dan pajak.",
    "Tulis angka polos tanpa titik ribuan. Tidak boleh 0.",
    "Benar : 1250000        Salah : Rp 1.250.000,-",
    "",
    "KHUSUS PRODUK BARANG",
    "Kolom Berat, Panjang, Lebar, dan Tinggi hanya berlaku bila kategorinya",
    "bertipe Barang. Untuk Jasa dan Digital, kosongkan — bagian itu memang",
    "tidak muncul di form. Berat wajib diisi untuk produk Barang.",
    "",
    "ATRIBUT KHUSUS KATEGORI",
    "Di form ada bagian 'Spesifikasi Produk > Informasi Utama' yang isinya",
    "berbeda tiap kategori — misalnya Satuan Pengukuran, Kode Produk, Lingkup",
    "Kegiatan, Lokasi Layanan, atau Kapasitas Tangki Septik.",
    "",
    "Tulis nama atributnya di kolom 'Atribut n', isinya di 'Nilai n' di",
    "sebelahnya. Nama harus sama persis dengan yang tertulis di form.",
    "Kalau semua produk satu kategori, isi sekali lalu tarik ke bawah.",
    "Slot yang tidak dipakai dikosongkan.",
    "",
    "YANG DIPERIKSA SAAT DIJALANKAN, BUKAN DI EXCEL",
    "Kode KBKI, Daftar Produk Sektoral, dan Satuan Pengukuran berbeda-beda",
    "tiap kategori, jadi tidak bisa dijadikan dropdown tetap di sini.",
    "Aplikasi memeriksanya setelah kategori dipilih di portal.",
    "",
    "YANG DIISI SENDIRI OLEH APLIKASI",
    "Kuantitas Desimal disimpulkan dari angka stok: 332,35 butuh desimal,",
    "1 tidak. Ongkir produk Barang selalu Standar.",
    "",
    "YANG BELUM DIDUKUNG",
    "Varian, harga grosir, harga zonasi, layanan tambahan, serta merek, SNI,",
    "dan TKDN. Kolomnya sengaja tidak disediakan: menyalakan saklarnya",
    "memunculkan isian lain di form yang belum bisa diisi aplikasi, jadi",
    "produk justru jadi tidak lengkap. Untuk sekarang diedit manual di portal.",
]


def _write_headers(ws, fields: tuple[Field, ...]) -> None:
    previous_group = None
    for index, fld in enumerate(fields, start=1):
        if fld.group != previous_group:
            cell = ws.cell(ROW_GROUP, index, fld.group)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="left")
            previous_group = fld.group
        ws.cell(ROW_GROUP, index).fill = _FILL_GROUP

        label = ws.cell(ROW_LABEL, index, fld.label)
        label.font = Font(bold=True, color="FFFFFF")
        label.fill = _FILL_LABEL
        label.alignment = Alignment(wrap_text=True, vertical="center")

        requirement = ws.cell(ROW_REQUIREMENT, index,
                              "(Wajib)" if fld.required else "(Opsional)")
        requirement.font = Font(size=9, italic=True)
        requirement.fill = _FILL_WAJIB if fld.required else _FILL_OPSIONAL

        hint = ws.cell(ROW_HINT, index, fld.hint)
        hint.font = Font(size=9, color="808080", italic=True)
        hint.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions[get_column_letter(index)].width = fld.width


def _rentang(sheet: str, kolom: str, baris_terakhir: int) -> str:
    return f"={quote_sheetname(sheet)}!${kolom}$2:${kolom}${baris_terakhir}"


def _pasang_dv(ws, index: int, formula: str, pesan: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = pesan
    dv.errorTitle = "Nilai tidak sah"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    letter = get_column_letter(index)
    dv.add(f"{letter}{FIRST_DATA_ROW}:{letter}{_MAX_VALIDATED_ROW}")


def _add_dropdowns(
    ws, fields: tuple[Field, ...], rujukan: dict[str, str], kategori_ref: str
) -> int:
    """Pasang dropdown. Daftar pendek ditulis di sel, yang panjang menunjuk sheet."""
    dipasang = 0
    for index, fld in enumerate(fields, start=1):
        if fld.key == "kategori" and kategori_ref:
            _pasang_dv(ws, index, kategori_ref,
                       "Salin jalur kategori dari sheet 'Daftar Kategori'.")
            dipasang += 1
            continue

        if not fld.options:
            continue

        inline = '"{}"'.format(",".join(fld.options))
        if len(inline) <= _BATAS_DAFTAR_INLINE:
            formula = inline
        elif fld.lookup and fld.lookup in rujukan:
            formula = rujukan[fld.lookup]
        else:
            continue  # daftar terlalu panjang dan tidak punya sheet rujukan

        _pasang_dv(ws, index, formula,
                   f"Pilih dari daftar yang tersedia untuk {fld.label}.")
        dipasang += 1
    return dipasang


def _write_reference_sheet(wb) -> dict[str, str]:
    """Sheet berisi daftar pilihan panjang; mengembalikan rumus rentang per daftar."""
    ws = wb.create_sheet(SHEET_PILIHAN)
    rumus: dict[str, str] = {}
    for offset, (nama, nilai) in enumerate(ref.DAFTAR_RUJUKAN.items()):
        index = offset + 1
        letter = get_column_letter(index)
        judul = ws.cell(1, index, nama)
        judul.font = Font(bold=True, color="FFFFFF")
        judul.fill = _FILL_LABEL
        for baris, isi in enumerate(nilai, start=2):
            ws.cell(baris, index, isi)
        ws.column_dimensions[letter].width = max(16, min(46, len(nama) + 6))
        rumus[nama] = _rentang(SHEET_PILIHAN, letter, len(nilai) + 1)
    ws.freeze_panes = "A2"
    return rumus


def _write_category_sheet(wb, catalog) -> tuple[int, str]:
    """Seluruh kategori portal sebagai sheet rujukan. Kolom E = jalur lengkap."""
    ws = wb.create_sheet(SHEET_KATEGORI)
    judul = ["Tipe Produk", "Kategori Level 1", "Kategori Level 2",
             "Kategori Level 3", "Jalur Kategori"]
    for index, teks in enumerate(judul, start=1):
        cell = ws.cell(1, index, teks)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _FILL_LABEL
    for lebar, huruf in zip((14, 34, 42, 52, 78), "ABCDE"):
        ws.column_dimensions[huruf].width = lebar

    baris = catalog.flatten()
    for offset, data in enumerate(baris, start=2):
        ws.cell(offset, 1, data["tipe_produk"])
        ws.cell(offset, 2, data["kategori_1"])
        ws.cell(offset, 3, data["kategori_2"])
        ws.cell(offset, 4, data["kategori_3"])
        ws.cell(offset, 5, join_category(
            data["kategori_1"], data["kategori_2"], data["kategori_3"]
        ))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(baris) + 1}"
    return len(baris), _rentang(SHEET_KATEGORI, "E", len(baris) + 1)


def create_template(
    path: str | Path, fields: tuple[Field, ...] | None = None, catalog=None
) -> Path:
    """Tulis template kosong siap diisi penyedia.

    Bila `catalog` diberikan, seluruh kategori portal ikut ditempel sebagai
    sheet rujukan sekaligus jadi sumber dropdown kolom Kategori.
    """
    fields = fields or all_fields()
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PRODUK
    _write_headers(ws, fields)

    kategori_ref = ""
    if catalog is not None and not catalog.kosong:
        _, kategori_ref = _write_category_sheet(wb, catalog)
    rujukan = _write_reference_sheet(wb)
    _add_dropdowns(ws, fields, rujukan, kategori_ref)

    ws.row_dimensions[ROW_LABEL].height = 32
    ws.row_dimensions[ROW_HINT].height = 28
    ws.freeze_panes = ws.cell(FIRST_DATA_ROW, 3)  # kunci judul + kolom kategori

    guide = wb.create_sheet(SHEET_PETUNJUK)
    guide.column_dimensions["A"].width = 78
    for offset, line in enumerate(PETUNJUK):
        cell = guide.cell(offset + 1, 1, line)
        if line and line.isupper():
            cell.font = Font(bold=True)

    wb.save(path)
    return path


def read_workbook(
    path: str | Path, fields: tuple[Field, ...] | None = None
) -> list[dict]:
    """Baca template terisi. Kunci hasil memakai `Field.key`, bukan judul kolom.

    Pencocokan lewat judul di baris 2, jadi penyedia boleh menggeser atau
    menyisipkan kolom tanpa merusak pembacaan.
    """
    fields = fields or all_fields()
    by_label = {f.label.strip().lower(): f.key for f in fields}

    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_PRODUK] if SHEET_PRODUK in wb.sheetnames else wb[wb.sheetnames[0]]

    columns: dict[int, str] = {}
    for index in range(1, ws.max_column + 1):
        label = ws.cell(ROW_LABEL, index).value
        if label is None or not str(label).strip():
            continue
        key = by_label.get(str(label).strip().lower())
        if key:  # kolom asing dibiarkan -- penyedia boleh menambah catatan
            columns[index] = key

    rows: list[dict] = []
    for excel_row in range(FIRST_DATA_ROW, ws.max_row + 1):
        record: dict = {"_row": excel_row}
        for index, key in columns.items():
            value = ws.cell(excel_row, index).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                record[key] = text
        if len(record) > 1:
            rows.append(record)
    return rows


def missing_columns(path: str | Path, fields: tuple[Field, ...] | None = None):
    """Judul kolom wajib yang tidak ditemukan di file."""
    fields = fields or all_fields()
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_PRODUK] if SHEET_PRODUK in wb.sheetnames else wb[wb.sheetnames[0]]
    present = {
        str(ws.cell(ROW_LABEL, i).value or "").strip().lower()
        for i in range(1, ws.max_column + 1)
    }
    return [f.label for f in fields if f.required and f.label.lower() not in present]


__all__ = [
    "FIRST_DATA_ROW",
    "ROW_LABEL",
    "SHEET_KATEGORI",
    "SHEET_PETUNJUK",
    "SHEET_PILIHAN",
    "SHEET_PRODUK",
    "create_template",
    "missing_columns",
    "read_workbook",
]
